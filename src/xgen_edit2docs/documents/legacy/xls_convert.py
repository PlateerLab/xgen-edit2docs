"""Excel 97 바이너리(.xls, BIFF8) → XLSX 변환 (reference_data/xlrd 대조).

'Workbook' 스트림은 (type 2B, len 2B) 레코드 나열이다. 워크북 전역부에서
SST(공유 문자열, CONTINUE 분절 규칙 포함)/FONT/XF/FORMAT/BOUNDSHEET 를
모으고, BOUNDSHEET 가 가리키는 시트 서브스트림에서 셀(NUMBER/RK/MULRK/
LABELSST/FORMULA 캐시값/BOOLERR)·행 높이(ROW)·열 너비(COLINFO)·병합
(MERGEDCELLS)을 읽어 openpyxl 워크북으로 재조립한다.

충실도 범위: 값·수식 캐시값·문자열(서식 run 은 평문화)·열너비/행높이·
병합·숫자서식·굵게/기울임/크기. 차트·이미지·조건부서식은 범위 밖.
BIFF5(.xls 구버전)는 SST 가 없어 지원하지 않는다 — 정직하게 거절한다.
"""

from __future__ import annotations

import io
import struct
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

from . import LegacyConvertError

# 레코드 타입
_R_BOF = 0x0809
_R_EOF = 0x000A
_R_SST = 0x00FC
_R_CONTINUE = 0x003C
_R_LABELSST = 0x00FD
_R_LABEL = 0x0204
_R_NUMBER = 0x0203
_R_RK = 0x027E
_R_MULRK = 0x00BD
_R_FORMULA = 0x0006
_R_STRING = 0x0207
_R_BOOLERR = 0x0205
_R_BOUNDSHEET = 0x0085
_R_ROW = 0x0208
_R_COLINFO = 0x007D
_R_MERGEDCELLS = 0x00E5
_R_FONT = 0x0031
_R_XF = 0x00E0
_R_FORMAT = 0x041E


def _iter_biff(data: bytes, start: int = 0):
    pos, n = start, len(data)
    while pos + 4 <= n:
        rtype, rlen = struct.unpack_from("<HH", data, pos)
        body_start = pos + 4
        if body_start + rlen > n:
            return
        yield rtype, body_start, rlen, pos
        pos = body_start + rlen


def _decode_short_unicode(data: bytes, pos: int) -> Tuple[str, int]:
    """BOUNDSHEET/FONT 이름용 짧은 유니코드 문자열 (1B cch + 1B flags)."""
    cch = data[pos]
    flags = data[pos + 1]
    pos += 2
    if flags & 0x01:
        s = data[pos:pos + cch * 2].decode("utf-16le", errors="replace")
        return s, pos + cch * 2
    return data[pos:pos + cch].decode("latin-1"), pos + cch


def _rk_value(raw: int) -> float:
    f_x100 = raw & 0x01
    f_int = raw & 0x02
    if f_int:
        # 상위 30비트가 부호 있는 정수 — 파이썬 int 는 무한 폭이라 32비트로
        # 재해석한 뒤 산술 시프트한다.
        v = float(struct.unpack("<i", struct.pack("<I", raw & 0xFFFFFFFF))[0] >> 2)
    else:
        # 상위 30비트가 IEEE754 double 의 상위 30비트 (하위 34비트 = 0).
        (v,) = struct.unpack("<d", b"\x00\x00\x00\x00" + struct.pack("<I", raw & 0xFFFFFFFC))
    return v / 100.0 if f_x100 else v


class _SstReader:
    """SST + CONTINUE — 분절 경계마다 문자열 조각 앞에 새 grbit 가 온다."""

    def __init__(self, fragments: List[bytes]):
        self.frags = fragments
        self.fi = 0
        self.pos = 0

    def _remaining(self) -> int:
        return len(self.frags[self.fi]) - self.pos

    def _advance_fragment(self) -> None:
        self.fi += 1
        self.pos = 0

    def read_exact(self, n: int) -> bytes:
        out = bytearray()
        while n > 0:
            if self.fi >= len(self.frags):
                raise LegacyConvertError("xls SST 가 중간에 끊겼습니다")
            take = min(n, self._remaining())
            if take == 0:
                self._advance_fragment()
                continue
            frag = self.frags[self.fi]
            out += frag[self.pos:self.pos + take]
            self.pos += take
            n -= take
        return bytes(out)

    def read_chars(self, cch: int, high: bool) -> str:
        """문자 데이터 — 분절 경계를 넘으면 그 지점에서 새 grbit 를 읽는다."""
        parts: List[str] = []
        remaining = cch
        while remaining > 0:
            if self.fi >= len(self.frags):
                raise LegacyConvertError("xls SST 문자열이 중간에 끊겼습니다")
            if self._remaining() == 0:
                self._advance_fragment()
                if remaining > 0:
                    grbit = self.read_exact(1)[0]
                    high = bool(grbit & 0x01)
                continue
            unit = 2 if high else 1
            fit = min(remaining, self._remaining() // unit)
            if fit == 0:
                # 홀수 바이트 걸침(이론상 없음) — 다음 분절로
                self._advance_fragment()
                if remaining > 0 and self.fi < len(self.frags):
                    grbit = self.read_exact(1)[0]
                    high = bool(grbit & 0x01)
                continue
            raw = self.read_exact(fit * unit)
            parts.append(raw.decode("utf-16le" if high else "latin-1",
                                    errors="replace"))
            remaining -= fit
        return "".join(parts)

    def read_string(self) -> str:
        cch = struct.unpack("<H", self.read_exact(2))[0]
        grbit = self.read_exact(1)[0]
        high = bool(grbit & 0x01)
        c_run = struct.unpack("<H", self.read_exact(2))[0] if grbit & 0x08 else 0
        cb_ext = struct.unpack("<I", self.read_exact(4))[0] if grbit & 0x04 else 0
        s = self.read_chars(cch, high)
        if c_run:
            self.read_exact(c_run * 4)
        if cb_ext:
            self.read_exact(cb_ext)
        return s


@dataclass
class _Font:
    size_pt: float = 10.0
    bold: bool = False
    italic: bool = False


def xls_to_xlsx(content: bytes) -> bytes:
    import olefile
    from openpyxl import Workbook
    from openpyxl.styles import Font as XlFont
    from openpyxl.styles.numbers import BUILTIN_FORMATS
    from openpyxl.utils import get_column_letter

    if not olefile.isOleFile(io.BytesIO(content)):
        raise LegacyConvertError("xls 가 아닙니다 (OLE 복합문서 아님)")
    ole = olefile.OleFileIO(io.BytesIO(content))
    try:
        stream_name = next((n for n in ("Workbook", "Book") if ole.exists(n)), None)
        if stream_name is None:
            raise LegacyConvertError("xls Workbook 스트림이 없습니다")
        data = ole.openstream(stream_name).read()
    finally:
        ole.close()

    # ── 워크북 전역부 ───────────────────────────────────────────
    sst: List[str] = []
    fonts: List[_Font] = []
    xf_font: List[int] = []
    xf_fmt: List[int] = []
    fmt_codes: Dict[int, str] = {}
    sheets: List[Tuple[str, int]] = []  # (이름, BOF 절대 오프셋)

    records = list(_iter_biff(data))
    if not records or records[0][0] != _R_BOF:
        raise LegacyConvertError("xls BOF 가 없습니다")
    (biff_ver,) = struct.unpack_from("<H", data, records[0][1])
    if biff_ver < 0x0600:
        raise LegacyConvertError("BIFF8(Excel 97+) 이전 .xls 는 지원하지 않습니다")

    i = 0
    depth = 0
    while i < len(records):
        rtype, bstart, rlen, _rpos = records[i]
        if rtype == _R_BOF:
            depth += 1
        elif rtype == _R_EOF:
            depth -= 1
            if depth == 0:
                break  # 전역부 끝 — 시트 서브스트림은 오프셋으로 접근
        elif depth == 1:
            if rtype == _R_SST:
                frags = [data[bstart + 8:bstart + rlen]]  # total/unique 건너뜀
                (unique,) = struct.unpack_from("<I", data, bstart + 4)
                j = i + 1
                while j < len(records) and records[j][0] == _R_CONTINUE:
                    cb_start, cb_len = records[j][1], records[j][2]
                    frags.append(data[cb_start:cb_start + cb_len])
                    j += 1
                reader = _SstReader(frags)
                for _ in range(unique):
                    sst.append(reader.read_string())
            elif rtype == _R_FONT:
                f = _Font()
                if rlen >= 8:
                    height, grbit, _icv, weight = struct.unpack_from(
                        "<HHHH", data, bstart)
                    f.size_pt = max(1.0, height / 20.0)
                    f.italic = bool(grbit & 0x0002)
                    f.bold = weight >= 600
                fonts.append(f)
            elif rtype == _R_XF and rlen >= 4:
                ifnt, ifmt = struct.unpack_from("<HH", data, bstart)
                xf_font.append(ifnt)
                xf_fmt.append(ifmt)
            elif rtype == _R_FORMAT and rlen >= 5:
                (ifmt,) = struct.unpack_from("<H", data, bstart)
                cch, flags = struct.unpack_from("<HB", data, bstart + 2)
                p = bstart + 5
                code = (data[p:p + cch * 2].decode("utf-16le", errors="replace")
                        if flags & 0x01 else data[p:p + cch].decode("latin-1"))
                fmt_codes[ifmt] = code
            elif rtype == _R_BOUNDSHEET and rlen >= 8:
                (bof_pos,) = struct.unpack_from("<I", data, bstart)
                name, _ = _decode_short_unicode(data, bstart + 6)
                sheets.append((name or f"Sheet{len(sheets) + 1}", bof_pos))
        i += 1

    if not sheets:
        raise LegacyConvertError("xls 에 시트가 없습니다")

    def font_of_xf(ixfe: int) -> Optional[_Font]:
        if not (0 <= ixfe < len(xf_font)):
            return None
        ifnt = xf_font[ixfe]
        # BIFF 규약: 폰트 인덱스 4는 존재하지 않는다 — 5 이상은 1 당긴다.
        if ifnt >= 4:
            ifnt -= 1
        return fonts[ifnt] if 0 <= ifnt < len(fonts) else None

    def fmt_of_xf(ixfe: int) -> Optional[str]:
        if not (0 <= ixfe < len(xf_fmt)):
            return None
        ifmt = xf_fmt[ixfe]
        if ifmt in fmt_codes:
            return fmt_codes[ifmt]
        return BUILTIN_FORMATS.get(ifmt)

    # ── 시트 서브스트림 ─────────────────────────────────────────
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, bof_pos in sheets:
        ws = wb.create_sheet(title=sheet_name[:31] or None)
        pending_string_cell = None  # FORMULA 문자열 결과 → 다음 STRING

        def put(r: int, c: int, ixfe: int, value) -> None:
            cell = ws.cell(row=r + 1, column=c + 1, value=value)
            f = font_of_xf(ixfe)
            if f and (f.bold or f.italic or abs(f.size_pt - 10.0) > 0.01):
                cell.font = XlFont(bold=f.bold, italic=f.italic, size=f.size_pt)
            code = fmt_of_xf(ixfe)
            if code and code.lower() != "general":
                cell.number_format = code

        for rtype, bstart, rlen, _rpos in _iter_biff(data, bof_pos):
            if rtype == _R_EOF:
                break
            if rtype == _R_LABELSST and rlen >= 10:
                r, c, ixfe, isst = struct.unpack_from("<HHHI", data, bstart)
                if 0 <= isst < len(sst):
                    put(r, c, ixfe, sst[isst])
            elif rtype == _R_NUMBER and rlen >= 14:
                r, c, ixfe = struct.unpack_from("<HHH", data, bstart)
                (v,) = struct.unpack_from("<d", data, bstart + 6)
                put(r, c, ixfe, v)
            elif rtype == _R_RK and rlen >= 10:
                r, c, ixfe, raw = struct.unpack_from("<HHHI", data, bstart)
                put(r, c, ixfe, _rk_value(raw))
            elif rtype == _R_MULRK and rlen >= 10:
                r, c_first = struct.unpack_from("<HH", data, bstart)
                count = (rlen - 6) // 6
                for k in range(count):
                    ixfe, raw = struct.unpack_from("<HI", data, bstart + 4 + k * 6)
                    put(r, c_first + k, ixfe, _rk_value(raw))
            elif rtype == _R_FORMULA and rlen >= 14:
                r, c, ixfe = struct.unpack_from("<HHH", data, bstart)
                result = data[bstart + 6:bstart + 14]
                if result[6:8] == b"\xff\xff":
                    kind = result[0]
                    if kind == 0:
                        pending_string_cell = (r, c, ixfe)
                    elif kind == 1:
                        put(r, c, ixfe, bool(result[2]))
                    elif kind == 2:
                        put(r, c, ixfe, "#ERR")
                else:
                    (v,) = struct.unpack("<d", result)
                    put(r, c, ixfe, v)
            elif rtype == _R_STRING and pending_string_cell is not None:
                reader = _SstReader([data[bstart:bstart + rlen]])
                try:
                    s = reader.read_string()
                except LegacyConvertError:
                    s = ""
                r, c, ixfe = pending_string_cell
                put(r, c, ixfe, s)
                pending_string_cell = None
            elif rtype == _R_BOOLERR and rlen >= 8:
                r, c, ixfe = struct.unpack_from("<HHH", data, bstart)
                v, is_err = data[bstart + 6], data[bstart + 7]
                put(r, c, ixfe, "#ERR" if is_err else bool(v))
            elif rtype == _R_ROW and rlen >= 16:
                r, = struct.unpack_from("<H", data, bstart)
                (miy,) = struct.unpack_from("<H", data, bstart + 6)
                if not (miy & 0x8000) and miy:
                    ws.row_dimensions[r + 1].height = (miy & 0x7FFF) / 20.0
            elif rtype == _R_COLINFO and rlen >= 6:
                c_first, c_last, width = struct.unpack_from("<HHH", data, bstart)
                for c in range(c_first, min(c_last, 255) + 1):
                    ws.column_dimensions[get_column_letter(c + 1)].width = width / 256.0
            elif rtype == _R_MERGEDCELLS and rlen >= 2:
                (count,) = struct.unpack_from("<H", data, bstart)
                for k in range(count):
                    off = bstart + 2 + k * 8
                    if off + 8 > bstart + rlen:
                        break
                    r1, r2, c1, c2 = struct.unpack_from("<4H", data, off)
                    if (r1, c1) != (r2, c2):
                        ws.merge_cells(start_row=r1 + 1, start_column=c1 + 1,
                                       end_row=r2 + 1, end_column=c2 + 1)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
